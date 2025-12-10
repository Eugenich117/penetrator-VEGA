import numpy as np
import time
from scipy.special import sindg, cosdg
import math as m
import matplotlib.pyplot as plt
from icecream import ic
import scipy
import concurrent.futures
import threading
import multiprocessing
from multiprocessing import Queue, Pipe
import random
import bisect
import sqlite3
import pandas as pd
from openpyxl import Workbook
from tkinter import filedialog
# мат модель из книжки воронцова упрощенная



# Cxa = 1.3#((2*L*r2*(1+r1/r2)/S))*(m.tan(Qk)/2)*(2*m.cos(0)**2*m.sin(Qk)**2+m.sin(0))
# Cya = 0#((2*L*r2*(1+r1/r2))/S)*m.pi*m.cos(0)*m.sin(0)*m.cos(Qk)*m.cos(Qk)
# Px = mass / Cxa * S
# K = Cya / Cxa
# ic(Cxa, Cya)



def find_closest_points_ro(x, xi):
    """
    Находит 4 ближайшие точки к xi в списке x.
    """
    closest_points = []
    for i in range(len(x)):
        if xi - 2 <= x[i] <= xi + 2:
            if i <= 2:
                closest_indices = list(range(4))
            elif i >= len(x) - 2:
                closest_indices = list(range(len(x) - 4, len(x)))
            else:
                closest_indices = list(range(i - 2, i + 2))
            closest_points = [x[idx] if 0 <= idx < len(x) else closest_points[-1] for idx in closest_indices]
            break
    return closest_points


def divided_diff_ro(x, y):
    """
    Вычисление разделённых разностей.
    """
    n = len(y)
    coef = [0] * n
    coef[0] = y[0]
    for j in range(1, n):
        for i in range(n - 1, j - 1, -1):
            if x[i] == x[i - j]:
                coef[i] = y[i]  # Просто присвоить значение y[i], чтобы избежать деления на ноль
            else:
                y[i] = (y[i] - y[i - 1]) / (x[i] - x[i - j])
                coef[i] = y[i]
    return coef


def newton_interpolation_ro(x, y, xi):
    """
    Интерполяция методом Ньютона.
    """
    closest_points = find_closest_points_ro(x, xi)
    x_interpolate = closest_points
    y_interpolate = [y[x.index(x_interpolate[0])], y[x.index(x_interpolate[1])], y[x.index(x_interpolate[2])],
                     y[x.index(x_interpolate[3])]]
    coef = divided_diff_ro(x_interpolate, y_interpolate)
    n = len(coef) - 1
    result = coef[n]
    for i in range(n - 1, -1, -1):
        result = result * (xi - x_interpolate[i]) + coef[i]
    return result


def find_closest_points(x, xi):
    """
    Находит две ближайшие точки к xi в списке x.
    """
    closest_points = []
    for i in range(len(x)):
        if x[i] >= xi:
            if i == 0:
                closest_points = [x[i], x[i + 1]]
            else:
                closest_points = [x[i - 1], x[i]]
            break
    return closest_points


def divided_diff(x, y):
    """
    Вычисление разделённых разностей.
    """
    n = len(y)
    coef = [0] * n
    coef[0] = y[0]
    for j in range(1, n):
        for i in range(n - 1, j - 1, -1):
            if x[i] == x[i - j]:
                coef[i] = y[i]  # Просто присвоить значение y[i], чтобы избежать деления на ноль
            else:
                y[i] = (y[i] - y[i - 1]) / (x[i] - x[i - j])
                coef[i] = y[i]
    return coef


def newton_interpolation(x, y, xi):
    """
    Интерполяция методом Ньютона.
    """
    closest_points = find_closest_points(x, xi)
    x_interpolate = closest_points
    y_interpolate = [y[x.index(x_interpolate[0])], y[x.index(x_interpolate[1])]]
    coef = divided_diff(x_interpolate, y_interpolate)
    n = len(coef) - 1
    result = coef[n]

    for i in range(n - 1, -1, -1):
        result = result * (xi - x_interpolate[i]) + coef[i]

    return result


def Get_ro(R): # В основной функции всё в метрах, в полиноме в километрах
    x = [130, 128, 126, 124, 122, 120, 118, 116, 114, 112, 110, 108, 106, 104, 102, 100, 98, 96, 94, 92, 90, 88, 86,
         84, 82, 80, 78, 76, 74, 72, 70, 68, 66, 64, 62, 60, 58, 56, 54, 52, 50, 48, 46, 44, 42, 40, 38, 36, 34, 32,
         31, 30, 29, 28, 27, 26, 25, 24, 23, 22, 21, 20, 19, 18, 17, 16, 15, 14, 13, 12, 11, 10, 9, 8, 7, 6, 5, 4, 3,
         2, 1, 0]
    y = [3.97200000e-08, 7.890 * 10 ** (-7), 1.35931821e-06, 1.77164551e-06, 2.30904567e-06, 3.00945751e-06, 3.92232801e-05,
         5.11210308e-05, 6.66277727e-05, 8.68382352e-05, 1.13179216e-04, 1.47510310e-04, 1.92255188e-04, 2.50572705e-04,
         3.26579903e-04, 1.347 * 10 ** (-4), 0.0002, 0.0004, 0.0007, 0.0012, 0.0019, 0.0031, 0.0049, 0.0077,
         0.0119, 0.0178, 0.0266, 0.0393, 0.0578, 0.0839, 0.1210, 0.1729, 0.2443, 0.3411, 0.4694, 0.6289,
         0.8183, 1.0320, 1.2840, 1.5940, 1.9670, 2.4260, 2.9850, 3.6460, 4.4040, 5.2760, 6.2740, 7.4200,
         8.7040, 9.4060, 10.1500, 10.9300, 11.7700, 12.6500, 13.5900, 14.5700, 15.6200, 16.7100, 17.8800,
         19.1100, 20.3900, 21.7400, 23.1800, 24.6800, 26.2700, 27.9500, 29.7400, 31.6000, 33.5400,
         35.5800, 37.7200, 39.9500, 42.2600, 44.7100, 47.2400, 49.8700, 52.6200, 55.4700, 58.4500, 61.5600,
         64.7900]
    ro = newton_interpolation_ro(x, y, R / 1000)
    return ro


def Cx(xi, V_sound):
    x = [0, 0, 0.2, 0.4, 0.6, 0.1, 1.2, 1.4, 1.8, 2, 2.2, 2.4, 2.6, 2.8, 3.2, 3.6, 4, 4.4, 4.8, 5.2, 6, 7, 8, 9, 10, 11,
         12, 13, 14, 15, 16, 17, 18, 19, 20, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38,
         39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66, 67]
    y = [0.75, 0.8, 0.9, 1.1, 1.3, 1.45, 1.52, 1.55, 1.6, 1.7, 1.8, 1.78, 1.75, 1.7, 1.65, 1.6, 1.55, 1.52, 1.52, 1.52,
         1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52,
         1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52,
         1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52,
         1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52, 1.52]
    return newton_interpolation(x, y, xi/V_sound)


def Cx_wind(xi, V_sound):
    x = [0, 0, 0.2, 0.4, 0.6, 0.1, 1.2, 1.4, 1.8, 2, 2.2, 2.4, 2.6, 2.8, 3.2, 3.6, 4, 4.4, 4.8, 5.2, 6, 7, 8, 9, 10, 11,
         12, 13, 14, 15, 16, 17, 18, 19, 20, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38,
         39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66, 67]
    y = [0.15, 0.15, 0.18, 0.3, 0.38, 0.81, 0.92, 0.97, 0.995, 0.991, 0.985, 0.98, 0.975, 0.97, 0.955, 0.935, 0.925,
         0.91, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9,
         0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9,
         0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9,
         0.9, 0.9, 0.9, 0.9]
    return newton_interpolation(x, y, xi/V_sound)


def v_sound(R):
    x = [130, 128, 126, 124, 122, 120, 118, 116, 114, 112, 110, 108, 106, 104, 102, 100, 98, 96, 94, 92, 90, 88, 86,
         84, 82, 80, 78, 76, 74, 72, 70, 68, 66, 64, 62, 60, 58, 56, 54, 52, 50, 48, 46, 44, 42, 40, 38, 36, 34, 32,
         31, 30, 29, 28, 27, 26, 25, 24, 23, 22, 21, 20, 19, 18, 17, 16, 15, 14, 13, 12, 11, 10, 9, 8, 7, 6, 5, 4, 3,
         2, 1, 0]
    y = [174, 176, 178, 180, 182, 185, 186, 187, 190, 193, 195, 196, 198, 199, 201, 203, 205, 206, 208.0, 208.0, 209.0,
         212.2, 215.4, 218.6, 221.8, 225.0, 228.2, 231.4, 234.6, 237.8, 241.0, 244.0, 247.0, 250.0, 253.0, 256.0, 263.2,
         270.4, 277.6, 284.8, 292.0, 296.8, 301.6, 306.4, 311.2, 316.0, 321.2, 326.4, 331.6, 336.8, 339.4, 342.0, 344.6,
         347.2, 349.8, 352.4, 355.0, 357.4, 359.8, 362.2, 364.6, 367.0, 369.4, 371.8, 374.2, 376.6, 379.0, 381.0, 383.0,
         385.0, 387.0, 389.0, 391.2, 393.4, 395.6, 397.8, 400.0, 402.0, 404.0, 406.0, 408.0, 410.0]
    return newton_interpolation_ro(x, y, R / 1000)


start_time = time.time()
r1 = 0.4
mass = 120
h = 125_000
mass_planet = 4.867 * 10 ** 24
Rb = 6_051_800
gravy_const = 6.67 * 10 ** (-11)
g = 8.869
dt = 0.01
tetta = -9
x = 0
y = 0
plotnost = []; CX = []

cToDeg = 180 / m.pi
cToRad = m.pi / 180

R = Rb + h
dV = 0


def sign(x):
    if x > 0:
        return 1
    elif x < 0:
        return -1
    else:
        return 0

def dV_func(initial):
    """
    ИСПРАВЛЕННОЕ уравнение для изменения скорости
    Основная ошибка: ветер должен влиять через ОТНОСИТЕЛЬНУЮ СКОРОСТЬ в аэродинамических силах,
    а не как отдельная сила
    """
    S = initial['S']
    R = initial['R']
    Cxa = initial['Cxa']
    ro = initial['ro']
    V = initial['V']
    tetta = initial['tetta']
    mass = initial['mass']
    V_wind = initial['V_wind']
    wind_angle = initial['wind_angle']
    Cn = initial['Cn']
    Fn = initial['Fn']

    # Гравитация
    g_local = 8.87 * (6051800 / R) ** 2

    # КРИТИЧЕСКОЕ ИСПРАВЛЕНИЕ: Ветер влияет через ОТНОСИТЕЛЬНУЮ СКОРОСТЬ
    # Проекция ветра на направление движения
    V_wind_parallel = V_wind * np.sin(wind_angle)

    # Относительная скорость для аэродинамических сил
    V_relative = V - V_wind_parallel

    # Сила сопротивления зависит от ОТНОСИТЕЛЬНОЙ скорости
    drag_force = 0.5 * ro * V_relative ** 2 * (Cxa * S + Cn * Fn)

    # Подъемная сила аэростата (проекция на траекторию)
    balloon_force = ro * 8.87 * np.sin(tetta)

    # Гравитация (проекция на траекторию)
    gravity_force = mass * g_local * np.sin(tetta)

    # Суммарная сила
    total_force = balloon_force - gravity_force - np.sign(V_relative) * drag_force

    # Эффективная масса
    effective_mass = mass + ro

    dV = total_force / effective_mass

    return dV, 'V'


def dtetta_func(initial):
    """
    ИСПРАВЛЕННОЕ уравнение для изменения угла
    Ветер не должен напрямую влиять на изменение угла
    """
    V = initial['V']
    tetta = initial['tetta']
    R = initial['R']
    ro = initial['ro']
    mass = initial['mass']

    # Гравитация
    g_local = 8.87 * (6051800 / R) ** 2

    # Подъемная сила аэростата (нормальная составляющая)
    lift_force = ro * 8.87 * np.cos(tetta)

    # Гравитационная составляющая
    gravity_component = g_local * np.cos(tetta)

    # Центробежная сила
    centrifugal = V / R if R > 6051800 else 0

    # Основное уравнение
    if V > 0.1:
        dtetta = (lift_force / (mass * V) - gravity_component / V + centrifugal)
    else:
        dtetta = centrifugal

    return dtetta, 'tetta'


def dL_func(initial):
    """
    ИСПРАВЛЕННОЕ уравнение для изменения дальности
    Ветер влияет на горизонтальное движение через снос
    """
    V = initial['V']
    tetta = initial['tetta']
    V_wind = initial['V_wind']
    wind_angle = initial['wind_angle']
    R = initial['R']

    # Горизонтальная скорость аппарата
    V_horizontal = V * np.cos(tetta)

    # Горизонтальная составляющая ветра
    V_wind_horizontal = V_wind * np.cos(wind_angle)

    # Суммарная горизонтальная скорость (аппарат + снос ветром)
    V_total_horizontal = V_horizontal + V_wind_horizontal

    dL = V_total_horizontal * (6051800 / R)

    return dL, 'L'


def dR_func(initial):
    """
    ИСПРАВЛЕННОЕ уравнение для изменения высоты
    Ветер слабо влияет на вертикальное движение
    """
    V = initial['V']
    tetta = initial['tetta']

    # Вертикальная скорость определяется только скоростью аппарата
    # Ветер практически не влияет на вертикальное движение
    dR = V * np.sin(tetta)

    return dR, 'R'


def wind(h, t, next_update_time, current_V_wind, current_wind_angle):
    """
    Улучшенная модель ветра с физическими ограничениями у поверхности
    """
    bounds = [0, 2000, 6000, 10000, 18000, 28000, 36000, 42000, 48000,
              55000, 61000, 68000, 76000, 85000, 94000, 100000, float('inf')]

    wind_ranges = [
        (0, 3),  # 0-2 км
        (3, 7),  # 2-6 км
        (7, 15),  # 6-10 км
        (15, 25),  # 10-18 км
        (25, 35),  # 18-28 км
        (30, 40),  # 28-36 км
        (35, 50),  # 36-42 км
        (45, 60),  # 42-48 км
        (55, 70),  # 48-55 км
        (65, 80),  # 55-61 км
        (70, 90),  # 61-68 км
        (85, 100),  # 68-76 км
        (75, 85),  # 76-85 км
        (60, 75),  # 85-94 км
        (10, 20),  # 94-100 км
        (0, 0)  # Выше 100 км
    ]

    #index = bisect.bisect_right(bounds, h) - 1

    if t >= next_update_time:
        #min_wind, max_wind = wind_ranges[index]

        #new_wind = np.random.uniform(min_wind, max_wind)
        new_wind = np.random.uniform(-3, 3)
        new_angle = np.random.uniform(0, np.pi / 2)
        wind_timer = np.random.uniform(2.0, 8.0)
        next_update_time = t + wind_timer
    else:
        new_wind = current_V_wind
        new_angle = current_wind_angle
        next_update_time = next_update_time

    return new_wind, new_angle, next_update_time


def runge_kutta_4(equations, initial, dt, dx):
    k1 = {key: 0 for key in dx}
    k2 = {key: 0 for key in dx}
    k3 = {key: 0 for key in dx}
    k4 = {key: 0 for key in dx}

    # k1
    for eq in equations:
        derivative, key = eq(initial)
        if key in dx:
            k1[key] = derivative

    # k2
    state_k2 = initial.copy()
    for key in dx:
        state_k2[key] += dt / 2 * k1[key]

    for eq in equations:
        derivative, key = eq(state_k2)
        if key in dx:
            k2[key] = derivative

    # k3
    state_k3 = initial.copy()
    for key in dx:
        state_k3[key] += dt / 2 * k2[key]

    for eq in equations:
        derivative, key = eq(state_k3)
        if key in dx:
            k3[key] = derivative

    # k4
    state_k4 = initial.copy()
    for key in dx:
        state_k4[key] += dt * k3[key]

    for eq in equations:
        derivative, key = eq(state_k4)
        if key in dx:
            k4[key] = derivative

    # итоговое обновление только для переменных из dx
    new_values = []
    for key in dx:
        new_val = initial[key] + (dt / 6) * (
                k1[key] + 2 * k2[key] + 2 * k3[key] + k4[key]
        )
        new_values.append(new_val)

    return new_values


#(i, napor, TETTA, X, Y, V_MOD, T, PX, nx, acceleration) # передача листов для результатов в явном виде в функцию
def compute_trajectory(i, equations, dx, pipe_conn):
    try:
        #print(f"поток {i} запущен")
        location = "start"
        t = 0
        d = 2.4
        S = (m.pi * d ** 2) / 4
        V, tetta, R, L = random.uniform(10_900, 11_100), random.uniform(-21, -17) * cToRad, Rb + h, 0
        #print(f'V = {V:.3f}, tetta = {tetta * cToDeg:.3f}')
        initial = {}
        initial['S'] = S
        initial['mass'] = 1855 #1755

        local_TETTA = []; local_X = []; local_Y = []; local_V_MOD = []; local_T = []; local_napor = []; local_nx = []
        local_PX = []; local_acceleration = []; local_v_wind = []; local_wind_angle = []
        next_update_time = -1
        V_wind = 0
        wind_angle = 0
        V_sound = v_sound(R - Rb)
        mach = V / V_sound
        while mach > 1.32:
            S, Cn, Fn, mass = 4.52, 0, 0, 1755 # 1855 #
            """этап 1 аэродинамическое торможение"""
            V_wind, wind_angle, next_update_time = wind(R - Rb, t, next_update_time, V_wind, wind_angle) #0, 0, 100000000#
            location = 'V_sound'
            V_sound = v_sound(R - Rb)
            location = 'ro'
            ro = Get_ro(R - Rb)
            Cxa = Cx(V, V_sound)
            Cxa_wind = Cx_wind(V, V_sound)
            Px = (mass / Cxa * S) * g

            initial.update(
                {'S': S, 'g': g, 'Cn': Cn, 'Fn': Fn, 'tetta': tetta, 'Cxa': Cxa, 'ro': ro, 'L': L, 'V': V, 'R': R,
                 'mass': mass, 'Cxa_wind': Cxa_wind, 'wind_angle': wind_angle, 'V_wind': V_wind})
            '''try:
                
            except Exception as e:
                print(f"Ошибка {e}")'''
            values = runge_kutta_4(equations, initial, dt, dx)
            V = values[0]
            L = values[1]
            tetta = values[2]
            R = values[3]
            t += dt

            mach = V / V_sound
            local_wind_angle.append(wind_angle)
            local_v_wind.append(V_wind)
            local_TETTA.append(tetta * cToDeg)
            local_X.append(L)
            local_Y.append(R - Rb)
            local_V_MOD.append(V)
            local_T.append(t)
            local_napor.append(0.5 * ro * V ** 2)
            local_nx.append((0.5 * S * Cxa * ro * V ** 2) / (mass * ((gravy_const * mass_planet) / R ** 2)))
            local_PX.append(Px)

        V_sound = v_sound(R - Rb)
        mach = V / V_sound
        #print(f' 1) V = {V:.3f}, tetta = {tetta * cToDeg:.3f}, L = {L:.3f}, H = {(R - Rb):.3f}, Mach={mach:.3f}, {t:.3f}')

        while mach > 0.74:
            """этап 2 спуск на парашюте увода """
            S, Cn, Fn, mass = 4.52, 0.65, 6, 1755 #1855 #
            V_wind, wind_angle, next_update_time = wind(R - Rb, t, next_update_time, V_wind, wind_angle) #0, 0, 100000000#
            location = 'V_sound'
            V_sound = v_sound(R - Rb)
            location = 'ro'
            ro = Get_ro(R - Rb)
            Cxa = Cx(V, V_sound)
            Cxa_wind = Cx_wind(V, V_sound)
            Px = (mass / Cxa * S) * g

            initial.update(
                {'S': S, 'g': g, 'Cn': Cn, 'Fn': Fn, 'tetta': tetta, 'Cxa': Cxa, 'ro': ro, 'L': L, 'V': V, 'R': R,
                 'mass': mass, 'Cxa_wind': Cxa_wind, 'wind_angle': wind_angle, 'V_wind': V_wind})
            values = runge_kutta_4(equations, initial, dt, dx)
            V = values[0]
            L = values[1]
            tetta = values[2]
            R = values[3]
            t += dt

            mach = V / V_sound
            local_wind_angle.append(wind_angle)
            local_v_wind.append(V_wind)
            local_TETTA.append(tetta * cToDeg)
            local_X.append(L)
            local_Y.append(R - Rb)
            local_V_MOD.append(V)
            local_T.append(t)
            local_napor.append(0.5 * ro * V ** 2)
            local_nx.append((0.5 * S * Cxa * ro * V ** 2) / (mass * ((gravy_const * mass_planet) / R ** 2)))
            local_PX.append(Px)

        V_sound = v_sound(R - Rb)
        mach = V / V_sound
        #print(f' 2) V = {V:.3f}, tetta = {tetta * cToDeg:.3f}, L = {L:.3f}, H = {(R - Rb):.3f}, Mach={mach:.3f}, {t:.3f}')

        while t <= 71:  # было 70 по циклограмме
            """третий этап спуск с верхней полусферой на парашюте увода"""
            S, Cn, Fn, mass = 4.155, 0.65, 6, 380 #480 #
            V_wind, wind_angle, next_update_time = wind(R - Rb, t, next_update_time, V_wind, wind_angle) #0, 0, 100000000#
            location = 'V_sound'
            V_sound = v_sound(R - Rb)
            location = 'ro'
            ro = Get_ro(R - Rb)
            Cxa = 1.28
            Cxa_wind = Cx_wind(V, V_sound)
            Px = (mass / Cxa * S) * g

            initial.update(
                {'S': S, 'g': g, 'Cn': Cn, 'Fn': Fn, 'tetta': tetta, 'Cxa': Cxa, 'ro': ro, 'L': L, 'V': V, 'R': R,
                 'mass': mass, 'Cxa_wind': Cxa_wind, 'wind_angle': wind_angle, 'V_wind': V_wind})
            values = runge_kutta_4(equations, initial, dt, dx)
            V = values[0]
            L = values[1]
            tetta = values[2]
            R = values[3]
            t += dt

            local_wind_angle.append(wind_angle)
            local_v_wind.append(V_wind)
            local_TETTA.append(tetta * cToDeg)
            local_X.append(L)
            local_Y.append(R - Rb)
            local_V_MOD.append(V)
            local_T.append(t)
            local_napor.append(0.5 * ro * V ** 2)
            local_nx.append((0.5 * S * Cxa * ro * V ** 2) / (mass * ((gravy_const * mass_planet) / R ** 2)))
            local_PX.append(Px)

        V_sound = v_sound(R - Rb)
        mach = V / V_sound
        #print(f' 3) V = {V:.3f}, tetta = {tetta * cToDeg:.3f}, L = {L:.3f}, H = {(R - Rb):.3f}, Mach={mach:.3f}, {t:.3f}')

        while t <= 231:  # while mach > 0.14: # # было 220 по циклограмме
            """четвертый этап спуск на стабилизирующем парашюте"""
            S, Cn, Fn, mass = 2.895, 0.78, 1.5, 125 # 225 #
            V_wind, wind_angle, next_update_time = wind(R - Rb, t, next_update_time, V_wind, wind_angle) #0, 0, 100000000#
            location = 'V_sound'
            V_sound = v_sound(R - Rb)
            location = 'ro'
            ro = Get_ro(R - Rb)
            Cxa = 0.58
            Cxa_wind = Cx_wind(V, V_sound)
            Px = (mass / Cxa * S) * g

            initial.update(
                {'S': S, 'g': g, 'Cn': Cn, 'Fn': Fn, 'tetta': tetta, 'Cxa': Cxa, 'ro': ro, 'L': L, 'V': V, 'R': R,
                 'mass': mass, 'Cxa_wind': Cxa_wind, 'wind_angle': wind_angle, 'V_wind': V_wind})
            values = runge_kutta_4(equations, initial, dt, dx)
            V = values[0]
            L = values[1]
            tetta = values[2]
            R = values[3]
            t += dt

            local_wind_angle.append(wind_angle)
            local_v_wind.append(V_wind)
            local_TETTA.append(tetta * cToDeg)
            local_X.append(L)
            local_Y.append(R - Rb)
            local_V_MOD.append(V)
            local_T.append(t)
            local_napor.append(0.5 * ro * V ** 2)
            local_nx.append((0.5 * S * Cxa * ro * V ** 2) / (mass * ((gravy_const * mass_planet) / R ** 2)))
            local_PX.append(Px)

        V_sound = v_sound(R - Rb)
        mach = V / V_sound
        #print(f' 4) V = {V:.3f}, tetta = {tetta * cToDeg:.3f}, L = {L:.3f}, H = {(R - Rb):.3f}, Mach={mach:.3f}, {t:.3f}')

        while t <= 400:  # while mach > 0.03: # было 400 по циклограмме
            """пятый этап спуск на парашюте ввода аэростата """
            S, Cn, Fn, mass = 2.895, 0.97, 35, 125 #225 #
            V_wind, wind_angle, next_update_time = wind(R - Rb, t, next_update_time, V_wind, wind_angle) #0, 0, 100000000#
            location = 'V_sound'
            V_sound = v_sound(R - Rb)
            location = 'ro'
            ro = Get_ro(R - Rb)
            Cxa = 0.58
            Cxa_wind = Cx_wind(V, V_sound)
            Px = (mass / Cxa * S) * g

            initial.update(
                {'S': S, 'g': g, 'Cn': Cn, 'Fn': Fn, 'tetta': tetta, 'Cxa': Cxa, 'ro': ro, 'L': L, 'V': V, 'R': R,
                 'mass': mass, 'Cxa_wind': Cxa_wind, 'wind_angle': wind_angle, 'V_wind': V_wind})
            values = runge_kutta_4(equations, initial, dt, dx)
            V = values[0]
            L = values[1]
            tetta = values[2]
            R = values[3]
            t += dt

            mach = V / V_sound
            local_wind_angle.append(wind_angle)
            local_v_wind.append(V_wind)
            local_TETTA.append(tetta * cToDeg)
            local_X.append(L)
            local_Y.append(R - Rb)
            local_V_MOD.append(V)
            local_T.append(t)
            local_napor.append(0.5 * ro * V ** 2)
            local_nx.append((0.5 * S * Cxa * ro * V ** 2) / (mass * ((gravy_const * mass_planet) / R ** 2)))
            local_PX.append(Px)
        print(f' 5) V = {V:.3f}, tetta = {tetta * cToDeg:.3f}, L = {L:.3f}, H = {(R - Rb):.3f}, Mach={mach:.3f}, {t:.3f}')

        location = 'compute_acceleration'
        for j in range(1, len(local_V_MOD)):
            derivative_value = (local_V_MOD[j] - local_V_MOD[j - 1]) / dt
            local_acceleration.append(derivative_value)
        location = 'pack_result'

        result = (i, local_TETTA, local_X, local_Y, local_V_MOD, local_T, local_napor, local_nx, local_PX,
                  local_acceleration, local_v_wind, local_wind_angle)

        try:
            pipe_conn.send(result)
        except (BrokenPipeError, EOFError):
            error_info = traceback.format_exc()
            if not pipe_conn.closed:
                pipe_conn.send(('error', (location, i, error_info)))
                print(f"[ОШИБКА] Процесс {i} на этапе {location}:\n{error_info}")
                print(f'V = {V:.3f}, tetta = {tetta * cToDeg:.3f}, L = {L:.3f}, H = {(R - Rb):.3f}, t = {t:.3f}')

    except Exception as e:
        try:
            # Передаем не только ошибку, но и место (location)
            if not pipe_conn.closed:
                pipe_conn.send(('error', (location, i, str(e))))
                print(f"[ОШИБКА] Процесс {i} на этапе {location}: {e}")
                print(f'V = {V:.3f}, tetta = {tetta * cToDeg:.3f}, L = {L:.3f}, H = {(R - Rb):.3f}, t = {t:.3f}')

        except Exception as e2:
            print(f"[ОШИБКА ПРИ ОТПРАВКЕ ОШИБКИ] Процесс {i}: {e2}")
            print(f'V = {V:.3f}, tetta = {tetta * cToDeg:.3f}, L = {L:.3f}, H = {(R - Rb):.3f}, t = {t:.3f}')


    finally:
        pipe_conn.close()
    gc.collect()



if __name__ == '__main__':
    iter = 100 #количество итераций
    dx = ['V', 'L', 'tetta', 'R']
    equations = [dV_func, dL_func, dtetta_func, dR_func]
    #with multiprocessing.Manager() as manager:

        # Создаем списки через менеджера
    '''manager = multiprocessing.Manager()
    acceleration = manager.list([[] for _ in range(5)])
    napor = manager.list([manager.list() for _ in range(5)])
    TETTA = manager.list([manager.list() for _ in range(5)])
    X = manager.list([manager.list() for _ in range(5)])
    Y = manager.list([manager.list() for _ in range(5)])
    T = manager.list([manager.list() for _ in range(5)])
    PX = manager.list([manager.list() for _ in range(5)])
    nx = manager.list([manager.list() for _ in range(5)])
    V_MOD = manager.list([manager.list() for _ in range(5)])'''
    '''manager = multiprocessing.Manager()
    acceleration = manager.list()
    napor = manager.list()
    TETTA = manager.list()
    X = manager.list()
    Y = manager.list()
    T = manager.list()
    PX = manager.list()
    nx = manager.list()
    V_MOD = manager.list()'''
    acceleration = ([[] for _ in range(iter)])
    napor = ([[] for _ in range(iter)])
    TETTA = ([[] for _ in range(iter)])
    X = ([[] for _ in range(iter)])
    Y = ([[] for _ in range(iter)])
    T = ([[] for _ in range(iter)])
    PX = ([[] for _ in range(iter)])
    nx = ([[] for _ in range(iter)])
    V_MOD = ([[] for _ in range(iter)])
    V_WIND = ([[] for _ in range(iter)])
    WIND_ANGLE = ([[] for _ in range(iter)])
    processes = []
    parent_conns = []
    child_conns = []

    tasks = []

    for i in range(iter):
        parent_conn, child_conn = multiprocessing.Pipe()  # Создаем пару для каждого процесса
        parent_conns.append(parent_conn)
        child_conns.append(child_conn)
        tasks.append((i, equations, dx, child_conn))  # Формируем задачи для передачи в пул


    # Функция обратного вызова, которая будет вызываться по завершении каждой задачи

    # Создаем пул процессов
    pool = multiprocessing.Pool(processes=30)

    # Запускаем задачи асинхронно с отслеживанием завершения
    for task in tasks:
        pool.apply_async(compute_trajectory, task)

    '''for i in range(iter):
        parent_conn, child_conn = Pipe()  # Создаем пару для каждого процесса
        parent_conns.append(parent_conn)
        child_conns.append(child_conn)
        p = multiprocessing.Process(target=compute_trajectory, args=(i, equations, dx, child_conn))
        p.start()
        processes.append(p)'''

    '''results = []
    while not queue.empty():
        results.append(queue.get())'''

    # Обработка результатов
    for i in range(iter):
        result = parent_conns[i].recv()
        (i, local_TETTA, local_X, local_Y, local_V_MOD, local_T, local_napor, local_nx, local_PX, local_acceleration,
         local_v_wind, local_wind_angle) = result
        TETTA[i] = local_TETTA
        X[i] = local_X
        Y[i] = local_Y
        V_MOD[i] = local_V_MOD
        T[i] = local_T
        napor[i] = local_napor
        nx[i] = local_nx
        PX[i] = local_PX
        acceleration[i] = local_acceleration
        WIND_ANGLE[i] = local_wind_angle
        V_WIND[i] = local_v_wind

    print("\n" + "=" * 80)
    print("ПРОВЕРКА КОРРЕКТНОСТИ ДАННЫХ ДЛЯ ВСЕХ ПЕРЕМЕННЫХ")
    print("=" * 80)

    # СОБИРАЕМ ПОСЛЕДНИЕ ЗНАЧЕНИЯ ДО ЛЮБЫХ МАНИПУЛЯЦИЙ
    last_values = {}

    # Переменные, которые НЕ делают pop() - берем последний элемент
    last_values['TETTA'] = [lst[-1] for lst in TETTA if len(lst) > 0]
    last_values['X'] = [lst[-1] for lst in X if len(lst) > 0]
    last_values['Y'] = [lst[-1] for lst in Y if len(lst) > 0]
    last_values['V_MOD'] = [lst[-1] for lst in V_MOD if len(lst) > 0]
    last_values['napor'] = [lst[-1] for lst in napor if len(lst) > 0]

    # Переменные, которые делают pop() в графиках - берем последний элемент ДО pop()
    last_values['T'] = [lst[-1] for lst in T if len(lst) > 0]  # Берем до T.pop()
    last_values['PX'] = [lst[-1] for lst in PX if len(lst) > 0]  # Берем до PX.pop()
    last_values['nx'] = [lst[-1] for lst in nx if len(lst) > 0]  # Берем до nx.pop()
    last_values['acceleration'] = [lst[-1] for lst in acceleration if len(lst) > 0]  # Берем до acceleration.pop()
    last_values['V_WIND'] = [lst[-1] for lst in V_WIND if len(lst) > 0]  # Берем до V_WIND.pop()
    last_values['WIND_ANGLE'] = [lst[-1] for lst in WIND_ANGLE if len(lst) > 0]  # Берем до WIND_ANGLE.pop()

    # ПРОВЕРКА ДЛИН МАССИВОВ
    print("Проверка длин исходных массивов:")
    for var_name in ['TETTA', 'X', 'Y', 'V_MOD', 'T', 'napor', 'PX', 'nx', 'acceleration', 'V_WIND', 'WIND_ANGLE']:
        if var_name in locals():
            lengths = [len(lst) for lst in eval(var_name) if len(lst) > 0]
            if lengths:
                print(f"  {var_name}: {len(lengths)} реализаций, длины: {set(lengths)}")
            else:
                print(f"  {var_name}: нет данных")

    # ПРОВЕРКА КОРРЕКТНОСТИ ПОСЛЕДНИХ ЗНАЧЕНИЙ
    print(f"\nПроверка последних значений для ключевых переменных:")
    print("=" * 80)

    # Проверяем диапазоны значений для каждой переменной
    check_variables = {
        'TETTA': (-95, -85, "°"),  # Ожидаемый диапазон угла
        'Y': (0, 100, "м"),  # Высота в конце (должна быть близка к 0)
        'V_MOD': (0, 10, "м/с"),  # Скорость в конце
        'X': (0, 50000, "м"),  # Дальность
        'napor': (0, 1000, "Па"),  # Скоростной напор
        'acceleration': (-50, 50, "м/с²")  # Ускорение
    }

    for var_name, (expected_min, expected_max, unit) in check_variables.items():
        if var_name in last_values and last_values[var_name]:
            values = last_values[var_name]
            actual_min = min(values)
            actual_max = max(values)
            actual_mean = np.mean(values)

            print(f"\n{var_name} ({unit}):")
            print(f"  Ожидаемый диапазон: [{expected_min}, {expected_max}]")
            print(f"  Фактический диапазон: [{actual_min:.2f}, {actual_max:.2f}]")
            print(f"  Мат. ожидание: {actual_mean:.2f}")

            # Проверяем, попадают ли значения в ожидаемый диапазон
            if actual_min < expected_min or actual_max > expected_max:
                print(f"  ⚠️  ВНИМАНИЕ: значения выходят за ожидаемый диапазон!")

            # Проверяем первые 5 значений
            print(f"  Первые 5 значений: {[f'{v:.2f}' for v in values[:5]]}")

    # ДЕТАЛЬНАЯ ПРОВЕРКА TETTA
    print(f"\n" + "=" * 80)
    print("ДЕТАЛЬНАЯ ПРОВЕРКА TETTA:")
    print("=" * 80)

    if 'TETTA' in last_values and last_values['TETTA']:
        tetta_values = last_values['TETTA']
        print(f"Количество реализаций: {len(tetta_values)}")
        print(f"Диапазон значений: от {min(tetta_values):.2f}° до {max(tetta_values):.2f}°")
        print(f"Мат. ожидание: {np.mean(tetta_values):.2f}°")
        print(f"Медиана: {np.median(tetta_values):.2f}°")

        # Анализ распределения
        below_85 = len([v for v in tetta_values if v < -85])
        below_80 = len([v for v in tetta_values if v < -80])

        print(f"Значения < -85°: {below_85}/{len(tetta_values)} ({below_85 / len(tetta_values) * 100:.1f}%)")
        print(f"Значения < -80°: {below_80}/{len(tetta_values)} ({below_80 / len(tetta_values) * 100:.1f}%)")

        # Выводим все значения TETTA для анализа
        print(f"\nВсе значения TETTA:")
        for i, val in enumerate(tetta_values):
            print(f"  Реализация {i:3d}: {val:7.2f}°")

    # ПРОВЕРКА СОГЛАСОВАННОСТИ ДАННЫХ
    print(f"\n" + "=" * 80)
    print("ПРОВЕРКА СОГЛАСОВАННОСТИ ДАННЫХ:")
    print("=" * 80)

    # Проверяем, что все переменные имеют одинаковое количество реализаций
    counts = {var: len(values) for var, values in last_values.items() if values}
    if len(set(counts.values())) > 1:
        print("⚠️  РАЗНОЕ КОЛИЧЕСТВО РЕАЛИЗАЦИЙ:")
        for var, count in counts.items():
            print(f"  {var}: {count}")
    else:
        print(f"✅ Все переменные имеют одинаковое количество реализаций: {list(counts.values())[0]}")

    # ПРОВЕРКА ФИЗИЧЕСКОЙ КОРРЕКТНОСТИ
    print(f"\n" + "=" * 80)
    print("ПРОВЕРКА ФИЗИЧЕСКОЙ КОРРЕКТНОСТИ:")
    print("=" * 80)

    # Проверяем физическую осмысленность конечных состояний
    if 'Y' in last_values and 'V_MOD' in last_values:
        y_values = last_values['Y']
        v_values = last_values['V_MOD']

        # Высота должна быть неотрицательной
        negative_height = len([y for y in y_values if y < 0])
        if negative_height > 0:
            print(f"⚠️  Обнаружена отрицательная высота в {negative_height} реализациях")

        # Скорость при посадке должна быть небольшой
        high_speed_landing = len([v for v in v_values if v > 20])  # Более 20 м/с при посадке
        if high_speed_landing > 0:
            print(f"⚠️  Высокая скорость посадки (>20 м/с) в {high_speed_landing} реализациях")

    # ВИЗУАЛИЗАЦИЯ РАСПРЕДЕЛЕНИЙ ВСЕХ ПЕРЕМЕННЫХ
    try:
        print(f"\nСоздание диагностических графиков...")

        variables_to_plot = [var for var in last_values.keys() if last_values[var]]
        n_vars = len(variables_to_plot)

        if n_vars > 0:
            n_cols = 3
            n_rows = (n_vars + n_cols - 1) // n_cols

            fig, axes = plt.subplots(n_rows, n_cols, figsize=(18, 5 * n_rows))
            if n_vars == 1:
                axes = [axes]
            else:
                axes = axes.flatten()

            for i, var_name in enumerate(variables_to_plot):
                if i < len(axes):
                    values = last_values[var_name]

                    axes[i].hist(values, bins=15, alpha=0.7, color='skyblue', edgecolor='black')
                    axes[i].axvline(np.mean(values), color='red', linestyle='-', linewidth=2,
                                    label=f'Среднее: {np.mean(values):.3f}')
                    axes[i].axvline(np.median(values), color='orange', linestyle='--', linewidth=2,
                                    label=f'Медиана: {np.median(values):.3f}')

                    axes[i].set_title(f'{var_name}\n(n={len(values)})', fontsize=12)
                    axes[i].set_xlabel('Значение')
                    axes[i].set_ylabel('Частота')
                    axes[i].legend()
                    axes[i].grid(True, alpha=0.3)

            # Скрываем пустые subplots
            for i in range(len(variables_to_plot), len(axes)):
                axes[i].set_visible(False)

            plt.tight_layout()
            plt.show()

    except Exception as e:
        print(f"Ошибка при построении диагностических графиков: {e}")

    print(f"\n" + "=" * 80)

    # Создаем новую книгу Excel
    wb = Workbook()
    step = 100  # Шаг для записи данных
    headers = ["PX", "Перегрузка", "Ускорение", "Скоростной напор", "Угол входа",
               "X", "Y", "Скорость", "Время"]  # Список заголовков для данных

    for variant in range(iter):
        # Создаем новый лист для каждого варианта
        if variant == 0:
            ws = wb.active
            ws.title = f"Вариант {variant + 1}"
        else:
            ws = wb.create_sheet(title=f"Вариант {variant + 1}")

        # Записываем заголовки в строку 2
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col_num, value=header)

        # Определяем максимальную длину среди массивов данных для текущего варианта
        max_len = max(len(PX[variant]), len(nx[variant]), len(acceleration[variant]),
                      len(napor[variant]), len(TETTA[variant]), len(X[variant]), len(Y[variant]),
                      len(V_MOD[variant]), len(T[variant]))
        start_row = 3  # Начинаем с первой строки после заголовков

        # Записываем данные массивов в столбцы с заданным шагом
        for row_num in range(0, max_len, step):
            if row_num < len(PX[variant]):
                ws.cell(row=start_row + row_num // step, column=1, value=PX[variant][row_num])
            if row_num < len(nx[variant]):
                ws.cell(row=start_row + row_num // step, column=2, value=nx[variant][row_num])
            if row_num < len(acceleration[variant]):
                ws.cell(row=start_row + row_num // step, column=3, value=acceleration[variant][row_num])
            if row_num < len(napor[variant]):
                ws.cell(row=start_row + row_num // step, column=4, value=napor[variant][row_num])
            if row_num < len(TETTA[variant]):
                ws.cell(row=start_row + row_num // step, column=5, value=TETTA[variant][row_num])
            if row_num < len(X[variant]):
                ws.cell(row=start_row + row_num // step, column=6, value=X[variant][row_num])
            if row_num < len(Y[variant]):
                ws.cell(row=start_row + row_num // step, column=7, value=Y[variant][row_num])
            if row_num < len(V_MOD[variant]):
                ws.cell(row=start_row + row_num // step, column=8, value=V_MOD[variant][row_num])
            if row_num < len(T[variant]):
                ws.cell(row=start_row + row_num // step, column=9, value=T[variant][row_num])

    # Выбираем путь для сохранения файла
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

    if file_path:
        wb.save(file_path)  # Если путь выбран, сохраняем файл по указанному пути
        print(f"Данные успешно сохранены в файл: {file_path}")
    else:
        print("Сохранение отменено пользователем")

    for i in range(iter):
        plt.plot(X[i], Y[i], label=f'Вариант {i+1}')
    plt.title('Траектории спуска зонда-пенетратора', fontsize=16, fontname='Times New Roman')
    plt.xlabel('Дальность, м', fontsize=16, fontname='Times New Roman')
    plt.ylabel('Высота, м', fontsize=16, fontname='Times New Roman')
    plt.subplots_adjust(left=0.15, right=0.95, top=0.9, bottom=0.15)
    #plt.legend()
    plt.grid(True)
    plt.show()

    for i in range(iter):
        plt.plot(T[i], Y[i], label=f'Вариант {i+1}')
    plt.title('Зависимость высоты от времени', fontsize=16, fontname='Times New Roman')
    plt.xlabel('Время, с', fontsize=16, fontname='Times New Roman')
    plt.ylabel('Высота, м', fontsize=16, fontname='Times New Roman')
    plt.subplots_adjust(left=0.15, right=0.95, top=0.9, bottom=0.15)
    #plt.legend()
    plt.grid(True)
    plt.show()

    for i in range(iter):
        plt.plot(T[i], V_MOD[i], label=f'Вариант {i+1}')
    plt.title('Зависимость модуля скорости от времени', fontsize=16, fontname='Times New Roman')
    plt.xlabel("Время, c", fontsize=16, fontname='Times New Roman')
    plt.ylabel(r'Скорость, $\frac{м}{с}$', fontsize=16, fontname='Times New Roman')
    plt.subplots_adjust(left=0.15, right=0.95, top=0.9, bottom=0.15)
    #plt.legend()
    plt.grid(True)
    plt.show()

    for i in range(iter):
        plt.plot(Y[i], V_MOD[i], label=f'Вариант {i+1}')
    plt.title('Зависимость модуля скорости от высоты', fontsize=16, fontname='Times New Roman')
    plt.xlabel("Высота, м", fontsize=16, fontname='Times New Roman')
    plt.ylabel(r'Скорость, $\frac{м}{с}$', fontsize=16, fontname='Times New Roman')
    plt.subplots_adjust(left=0.15, right=0.95, top=0.9, bottom=0.15)
    #plt.legend()
    plt.grid(True)
    plt.show()

    for i in range(iter):
        plt.plot(T[i], TETTA[i], label=f'Вариант {i+1}')
    plt.title('Зависимость траекторного угла от времени', fontsize=16, fontname='Times New Roman')
    plt.xlabel('Время, c', fontsize=16, fontname='Times New Roman')
    plt.ylabel('Траекторный угол, град', fontsize=16, fontname='Times New Roman')
    plt.subplots_adjust(left=0.15, right=0.95, top=0.9, bottom=0.15)
    #plt.legend()
    plt.grid(True)
    plt.show()

    for i in range(iter):
        plt.plot(T[i], napor[i], label=f'Вариант {i+1}')
    plt.title('Зависимость скоростного напора от времени', fontsize=16, fontname='Times New Roman')
    plt.xlabel('Время, с', fontsize=16, fontname='Times New Roman')
    plt.ylabel(r'Скоростной напор, $\frac{\mathrm{кг}}{\mathrm{м} \cdot \mathrm{с}^{2}}$', fontsize=16, fontname='Times New Roman')
    plt.subplots_adjust(left=0.25, right=0.95, top=0.9, bottom=0.15)
    #plt.legend()
    plt.grid(True)
    plt.show()

    #T.pop()# Убираем последний элемент из списка времени
    for i in range(iter):
        T[i].pop()
        plt.plot(T[i], acceleration[i], label=f'Вариант {i+1}')
    plt.title('Зависимость ускорения от времени', fontsize=16, fontname='Times New Roman')
    plt.xlabel('Время, с', fontsize=16, fontname='Times New Roman')
    plt.ylabel('Ускорение м$^2$', fontsize=16, fontname='Times New Roman')
    plt.subplots_adjust(left=0.15, right=0.95, top=0.9, bottom=0.15)
    #plt.legend()
    plt.grid(True)
    plt.show()

    for i in range(iter):
        nx[i].pop()
        plt.plot(T[i], nx[i], label=f'Вариант {i + 1}')
    plt.title('Зависимость перегрузки от времени', fontsize=16, fontname='Times New Roman')
    plt.xlabel('Время, с', fontsize=16, fontname='Times New Roman')
    plt.ylabel('Перегрузка, g', fontsize=16, fontname='Times New Roman')
    plt.subplots_adjust(left=0.15, right=0.95, top=0.9, bottom=0.15)
    #plt.legend()
    plt.grid(True)
    plt.show()

    for i in range(iter):
        PX[i].pop()
        plt.plot(T[i], PX[i], label=f'Вариант {i + 1}')
    plt.title('Баллистический параметр', fontsize=16, fontname='Times New Roman')
    #plt.title('Зависимость давления на мидель от времени', fontsize=16, fontname='Times New Roman')
    plt.xlabel('Время, с', fontsize=16, fontname='Times New Roman')
    plt.ylabel(r'Px, $\frac{кг}{м^2}$', fontsize=16, fontname='Times New Roman')
    plt.subplots_adjust(left=0.15, right=0.95, top=0.9, bottom=0.15)
    #plt.legend()
    plt.grid(True)
    plt.show()

    for i in range(iter):
        V_WIND[i].pop()
        plt.plot(T[i], V_WIND[i], label=f'Вариант {i + 1}')
    plt.title('Скорость ветра от времени', fontsize=16, fontname='Times New Roman')
    plt.xlabel('Время, с', fontsize=16, fontname='Times New Roman')
    plt.ylabel(r'м/с', fontsize=16, fontname='Times New Roman')
    plt.subplots_adjust(left=0.15, right=0.95, top=0.9, bottom=0.15)
    #plt.legend()
    plt.grid(True)
    plt.show()

    for i in range(iter):
        WIND_ANGLE[i].pop()
        plt.plot(T[i], WIND_ANGLE[i], label=f'Вариант {i + 1}')
    plt.title('Угол действия ветра от времени', fontsize=16, fontname='Times New Roman')
    plt.xlabel('Время, с', fontsize=16, fontname='Times New Roman')
    plt.ylabel(r'Рад/с', fontsize=16, fontname='Times New Roman')
    plt.subplots_adjust(left=0.15, right=0.95, top=0.9, bottom=0.15)
    #plt.legend()
    plt.grid(True)
    plt.show()

    for i in range(iter):
        T1 = T[i][30000:]
        Y1 = Y[i][30000:]
        Y1.pop()
        plt.plot(T1, Y1, label=f'Вариант {i + 1}')
    plt.title('Зависимость высоты от времени', fontsize=16, fontname='Times New Roman')
    plt.xlabel('Время, с', fontsize=16, fontname='Times New Roman')
    plt.ylabel('Высота', fontsize=16, fontname='Times New Roman')
    plt.subplots_adjust(left=0.15, right=0.95, top=0.9, bottom=0.15)
    #plt.legend()
    plt.grid(True)
    plt.show()

    for i in range(iter):
        T[i] = T[i][:3000]
        napor[i] = napor[i][:3000]
        plt.plot(T[i], napor[i], label=f'Вариант {i + 1}')
    plt.title('Зависимость скоростного напора от времени', fontsize=16, fontname='Times New Roman')
    plt.xlabel('Время, с', fontsize=16, fontname='Times New Roman')
    plt.ylabel(r'Скоростной напор', fontsize=16, fontname='Times New Roman')
    plt.subplots_adjust(left=0.15, right=0.95, top=0.9, bottom=0.15)
    # plt.legend()
    plt.grid(True)
    plt.show()

    for i in range(iter):
        # Qk[i].pop()
        T[i] = T[i][:3000]
        slice_V_MOD = V_MOD[i][:3000]
        plt.plot(T[i], slice_V_MOD)
    plt.title('Зависимость модуля скорости от времени')
    plt.xlabel('Время, с')
    plt.ylabel('Q, КВт/м^2')
    plt.grid(True)
    plt.show()

    for i in range(iter):
        T[i] = T[i][:3000]
        TETTA[i] = TETTA[i][:3000]
        plt.plot(T[i], TETTA[i])
    plt.title('Зависимость угла входа от времени')
    plt.xlabel('Время, с')
    plt.ylabel('Q, KДж/м^2')
    plt.grid(True)
    plt.show()

    for i in range(iter):
        y_slice = Y[i][30000:]
        V_MOD[i] = V_MOD[i][30000:]
        plt.plot(y_slice, V_MOD[i])
    plt.title('Зависимость модуля скорости от высоты')
    plt.xlabel('Время, с')
    plt.ylabel('T, K')
    plt.grid(True)
    plt.show()

    for i in range(iter):
        T[i] = T[i][:3000]
        acceleration[i] = acceleration[i][:3000]
        plt.plot(T[i], acceleration[i], label=f'Вариант {i + 1}')
    plt.title('Зависимость ускорения от времени', fontsize=16, fontname='Times New Roman')
    plt.xlabel('Время, с', fontsize=16, fontname='Times New Roman')
    plt.ylabel('Ускорение', fontsize=16, fontname='Times New Roman')
    plt.subplots_adjust(left=0.15, right=0.95, top=0.9, bottom=0.15)
    #plt.legend()
    plt.grid(True)
    plt.show()

    for i in range(iter):
        T[i] = T[i][:3000]
        nx[i] = nx[i][:3000]
        plt.plot(T[i], nx[i], label=f'Вариант {i + 1}')
    plt.title('Зависимость перегрузки от времени', fontsize=16, fontname='Times New Roman')
    plt.xlabel('Время, с', fontsize=16, fontname='Times New Roman')
    plt.ylabel('Перегрузка, g', fontsize=16, fontname='Times New Roman')
    plt.subplots_adjust(left=0.15, right=0.95, top=0.9, bottom=0.15)
    #plt.legend()
    plt.grid(True)
    plt.show()

    print("\n" + "=" * 80)
    print("ФИНАЛЬНЫЕ СТАТИСТИКИ ПОСЛЕДНИХ ЗНАЧЕНИЙ")
    print("=" * 80)


    def calculate_final_stats(values, variable_name):
        """Рассчитывает статистики для последних значений"""
        if not values:
            return None

        values_array = np.array(values)

        return {
            'mean': np.mean(values_array),
            'variance': np.var(values_array, ddof=1),
            'std': np.std(values_array, ddof=1),
            'min': np.min(values_array),
            'max': np.max(values_array),
            'median': np.median(values_array),
            'count': len(values)
        }


    # Выводим финальные статистики
    print(
        f"{'Переменная':<15} {'Мат.ожидание':<15} {'Дисперсия':<15} {'СКО':<15} {'Min':<12} {'Max':<12} {'Реализаций':<10}")
    print("-" * 100)

    final_stats = {}
    for var_name in last_values.keys():
        stats = calculate_final_stats(last_values[var_name], var_name)
        if stats:
            final_stats[var_name] = stats
            print(f"{var_name:<15} {stats['mean']:<15.4f} {stats['variance']:<15.4f} "
                  f"{stats['std']:<15.4f} {stats['min']:<12.4f} {stats['max']:<12.4f} {stats['count']:<10}")

    print(f"\n✅ Анализ завершен для {len(final_stats)} переменных")

    data = {
        "acceleration": acceleration,
        "napor": napor,
        "TETTA": TETTA,
        "X": X,
        "Y": Y,
        "T": T,
        "PX": PX,
        "nx": nx,
        "V_MOD": V_MOD,
    }

    # Перебираем каждый массив и находим min и max
    for key, values in data.items():
        all_values = np.concatenate(values)  # Объединяем все списки в один массив
        min_val = np.min(all_values)
        max_val = np.max(all_values)
        print(f"{key}: min = {min_val}, max = {max_val}")

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(elapsed_time)