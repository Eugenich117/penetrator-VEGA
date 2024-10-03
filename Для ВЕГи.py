import numpy as np
import time
from scipy.special import sindg, cosdg
import math as m
import matplotlib.pyplot as plt
from icecream import ic
import scipy
from openpyxl import Workbook
from tkinter import filedialog

# мат модель из книжки воронцова упрощенная

r1 = 0.4
d = 2.4 #"""диаметр корпуса аппарата в начальный момент"""
#L = 0.53
"""задание констант и начальных условий"""
h = 125_000
mass_planet = 4.867*10**24
Rb = 6_051_800
gravy_const = 6.67*10**(-11)
g = 8.87
"""дистанция, угол входа, скорость gamma не """
L = 0
tetta = -19  # * (m.pi / 180)
V = np.float64(11_000)  # Используем тип данных float64
gamma = np.float64(0)  # Используем тип данных float64
dR = 0
dt = 0.01
dtetta = 0
t = 0.0


# Cxa = 1.3#((2*L*r2*(1+r1/r2)/S))*(m.tan(Qk)/2)*(2*m.cos(0)**2*m.sin(Qk)**2+m.sin(0))
# Cya = 0#((2*L*r2*(1+r1/r2))/S)*m.pi*m.cos(0)*m.sin(0)*m.cos(Qk)*m.cos(Qk)
# Px = mass / Cxa * S
# K = Cya / Cxa
# ic(Cxa, Cya)



def find_closest_points_ro(x, xi):
    """
    Находит 4 ближайших точек к xi в списке x.
    """
    closest_points = []
    for i in range(len(x)):
        if xi - 2 <= x[i] <= xi + 2:
            if i <= 2:
                closest_indices = list(range(4))
            elif i >= len(x) - 2:
                closest_indices = list(range(len(x) - 4, len(x)))
            else:
                closest_indices = list(range(i - 2, i + 2))
            closest_points = [x[idx] if 0 <= idx < len(x) else closest_points[-1] for idx in closest_indices]
            break
    return closest_points


def divided_diff_ro(x, y):
    """
    Вычисление разделённых разностей.
    """
    n = len(y)
    coef = [0] * n
    coef[0] = y[0]
    for j in range(1, n):
        for i in range(n - 1, j - 1, -1):
            if x[i] == x[i - j]:
                coef[i] = y[i]  # Просто присвоить значение y[i], чтобы избежать деления на ноль
            else:
                y[i] = (y[i] - y[i - 1]) / (x[i] - x[i - j])
                coef[i] = y[i]
    return coef


def newton_interpolation_ro(x, y, xi):
    """
    Интерполяция методом Ньютона.
    """
    closest_points = find_closest_points_ro(x, xi)
    x_interpolate = closest_points
    y_interpolate = [y[x.index(x_interpolate[0])], y[x.index(x_interpolate[1])], y[x.index(x_interpolate[2])],
                     y[x.index(x_interpolate[3])]]
    coef = divided_diff_ro(x_interpolate, y_interpolate)
    n = len(coef) - 1
    result = coef[n]
    for i in range(n - 1, -1, -1):
        result = result * (xi - x_interpolate[i]) + coef[i]
    return result


def find_closest_points(x, xi):
    """
    Находит две ближайшие точки к xi в списке x.
    """
    closest_points = []
    for i in range(len(x)):
        if x[i] >= xi:
            if i == 0:
                closest_points = [x[i], x[i + 1]]
            else:
                closest_points = [x[i - 1], x[i]]
            break
    return closest_points


def divided_diff(x, y):
    """
    Вычисление разделённых разностей.
    """
    n = len(y)
    coef = [0] * n
    coef[0] = y[0]
    for j in range(1, n):
        for i in range(n - 1, j - 1, -1):
            if x[i] == x[i - j]:
                coef[i] = y[i]
            else:
                y[i] = (y[i] - y[i - 1]) / (x[i] - x[i - j])
                coef[i] = y[i]
    return coef


def newton_interpolation(x, y, xi):
    """
    Интерполяция методом Ньютона. Функция, которая собирает все и вычисляет 1 значение
    """
    closest_points = find_closest_points(x, xi)
    x_interpolate = closest_points
    y_interpolate = [y[x.index(x_interpolate[0])], y[x.index(x_interpolate[1])]]
    coef = divided_diff(x_interpolate, y_interpolate)
    n = len(coef) - 1
    result = coef[n]

    for i in range(n - 1, -1, -1):
        result = result * (xi - x_interpolate[i]) + coef[i]

    return result


def Get_ro(R):
    """функция для задания массивов для интерполяции и вызова функции интерполяции"""
    x = [130, 128, 126, 124, 122, 120, 118, 116, 114, 112, 110, 108, 106, 104, 102, 100, 98, 96, 94, 92, 90, 88, 86,
         84, 82, 80, 78, 76, 74, 72, 70, 68, 66, 64, 62, 60, 58, 56, 54, 52, 50, 48, 46, 44, 42, 40, 38, 36, 34, 32,
         31, 30, 29, 28, 27, 26, 25, 24, 23, 22, 21, 20, 19, 18, 17, 16, 15, 14, 13, 12, 11, 10, 9, 8, 7, 6, 5, 4, 3,
         2, 1, 0]
    y = [3.97200000e-08, 7.890 * 10 ** (-7), 1.35931821e-06, 1.77164551e-06, 2.30904567e-06, 3.00945751e-06, 3.92232801e-05,
         5.11210308e-05, 6.66277727e-05, 8.68382352e-05, 1.13179216e-04, 1.47510310e-04, 1.92255188e-04, 2.50572705e-04,
         3.26579903e-04, 1.347 * 10 ** (-4), 0.0002, 0.0004, 0.0007, 0.0012, 0.0019, 0.0031, 0.0049, 0.0077,
         0.0119, 0.0178, 0.0266, 0.0393, 0.0578, 0.0839, 0.1210, 0.1729, 0.2443, 0.3411, 0.4694, 0.6289,
         0.8183, 1.0320, 1.2840, 1.5940, 1.9670, 2.4260, 2.9850, 3.6460, 4.4040, 5.2760, 6.2740, 7.4200,
         8.7040, 9.4060, 10.1500, 10.9300, 11.7700, 12.6500, 13.5900, 14.5700, 15.6200, 16.7100, 17.8800,
         19.1100, 20.3900, 21.7400, 23.1800, 24.6800, 26.2700, 27.9500, 29.7400, 31.6000, 33.5400,
         35.5800, 37.7200, 39.9500, 42.2600, 44.7100, 47.2400, 49.8700, 52.6200, 55.4700, 58.4500, 61.5600,
         64.7900]
    ro = newton_interpolation_ro(x, y, R /1000)
    return ro


def Cx(xi, V_sound):
    """функиця задания массивов и вычисления интерполированных значений для лобового сопротивления аппарата, когда он в
    сферическом корпусе для начала полета"""
    x = [0, 0, 0.2, 0.4, 0.6, 0.1, 1.2, 1.4, 1.8, 2, 2.2, 2.4, 2.6, 2.8, 3.2, 3.6, 4, 4.4, 4.8, 5.2, 6, 7, 8, 9, 10, 11,
         12, 13, 14, 15, 16, 17, 18, 19, 20, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38,
         39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66, 67]
    y = [0.15, 0.15, 0.18, 0.3, 0.38, 0.81, 0.92, 0.97, 0.995, 0.991, 0.985, 0.98, 0.975, 0.97, 0.955, 0.935, 0.925,
         0.91, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9,
         0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9,
         0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9, 0.9,
         0.9, 0.9, 0.9, 0.9]
    return newton_interpolation(x, y, xi/V_sound)


def v_sound(R):
    """вычисление интерполированных значений для скорсти звука """
    x = [130, 128, 126, 124, 122, 120, 118, 116, 114, 112, 110, 108, 106, 104, 102, 100, 98, 96, 94, 92, 90, 88, 86,
         84, 82, 80, 78, 76, 74, 72, 70, 68, 66, 64, 62, 60, 58, 56, 54, 52, 50, 48, 46, 44, 42, 40, 38, 36, 34, 32,
         31, 30, 29, 28, 27, 26, 25, 24, 23, 22, 21, 20, 19, 18, 17, 16, 15, 14, 13, 12, 11, 10, 9, 8, 7, 6, 5, 4, 3,
         2, 1, 0]
    y = [174, 176, 178, 180, 182, 185, 186, 187, 190, 193, 195, 196, 198, 199, 201, 203, 205, 206, 208.0, 208.0, 209.0,
         212.2, 215.4, 218.6, 221.8, 225.0, 228.2, 231.4, 234.6, 237.8, 241.0, 244.0, 247.0, 250.0, 253.0, 256.0, 263.2,
         270.4, 277.6, 284.8, 292.0, 296.8, 301.6, 306.4, 311.2, 316.0, 321.2, 326.4, 331.6, 336.8, 339.4, 342.0, 344.6,
         347.2, 349.8, 352.4, 355.0, 357.4, 359.8, 362.2, 364.6, 367.0, 369.4, 371.8, 374.2, 376.6, 379.0, 381.0, 383.0,
         385.0, 387.0, 389.0, 391.2, 393.4, 395.6, 397.8, 400.0, 402.0, 404.0, 406.0, 408.0, 410.0]
    return newton_interpolation_ro(x, y, R /1000)


start_time = time.time()
"""создаем массивы для всех величин, которые считаются в мат модели"""
x = 0
y = 0
PX = []
nx = []
plotnost = []
acceleration = []
napor = []
TETTA = []
CX = []
X = []
Y = []
V_MOD = []
T = []
R = Rb + h #"""в эту переменную суммируются радиус планеты и высота входа в атмосферу, далее мат модель работает только с ней"""
dV = 0
"""(gravy_const*mass_planet)/R**2) - это вычисление ускорения свободного падения на конкретной высоте"""
"""функции scipy.special.sindg() считают тригонометрические величины сразу в градусах"""

def dV_func(initial):
    """вычисление скорости"""
    S = initial['S']
    R = initial['R']
    Cxa = initial['Cxa']
    ro = initial['ro']
    V = initial['V']
    tetta = initial['tetta']
    Cn = initial['Cn']
    Fn = initial['Fn']
    #dV = ((-1 / (2 * Px)) * Cxa * ro * V ** 2 - ((gravy_const*mass_planet)/R**2) * scipy.special.sindg(tetta)) * dt # ОСНОВНАЯ МОДЕЛЬ КОСЕНКОВОЙ
    dV = ((-mass * ((gravy_const*mass_planet) / R ** 2) * m.sin(tetta) - (0.5 * ro * V ** 2 * (Cxa * S + Cn * Fn)))) / mass
    return dV, 'V'

def dL_func(initial):
    """вычисляеем длину траектории"""
    V = initial['V']
    tetta = initial['tetta']
    dL = V * Rb / R * m.cos(tetta)
    return dL, 'L'

def dtetta_func(initial):
    """вычисление угла наклона таектории"""
    V = initial['V']
    tetta = initial['tetta']
    R = initial['R']
    g=((gravy_const * mass_planet) / R ** 2)
    dtetta = -((m.cos(tetta)) * ((g / V) - (V / R)))
    #dtetta = ((-g * ((scipy.special.cosdg(tetta))/V)+(V/R))) * dt #был +
    #dtetta = ( ((V ** 2 - ((gravy_const*mass_planet)/R**2) * R) / (V * R)) * scipy.special.cosdg(tetta)) * dt
    return dtetta, 'tetta'

def dR_func(initial):
    """вычисление высоты"""
    V = initial['V']
    tetta = initial['tetta']
    dR = (V * m.sin(tetta))
    return dR, 'R'


def runge_kutta_4(equations, initial, dt, dx):
    '''equations - это список названий функций с уравнениями для системы
    initial это переменные с начальными условиями
    dx - это список переменных, которые будут использованы для интегрирования уравнения'''
    k1 = {key: 0 for key in initial.keys()}
    k2 = {key: 0 for key in initial.keys()}
    k3 = {key: 0 for key in initial.keys()}
    k4 = {key: 0 for key in initial.keys()}

    derivatives_1 = {key: initial[key] for key in initial}
    derivatives_2 = {key: initial[key] for key in initial}
    derivatives_3 = {key: initial[key] for key in initial}
    derivatives_4 = {key: initial[key] for key in initial}

    new_values = [0] * len(equations)

    for i, eq in enumerate(equations):
        derivative, key = eq(initial)
        k1[key] += derivative
        derivatives_1[key] = initial[key] + derivative * dt / 2
        derivatives_1[dx[i]] += dt / 2
        # derivatives_1 = {key: value / 2 for key, value in derivatives_1.items()}

    for i, eq in enumerate(equations):
        derivative, key = eq(derivatives_1)
        k2[key] += derivative
        derivatives_2[key] = initial[key] + derivative * dt / 2
        derivatives_2[dx[i]] += dt / 2
        # derivatives_2 = {key: value / 2 for key, value in derivatives_2.items()}

    for i, eq in enumerate(equations):
        derivative, key = eq(derivatives_2)
        k3[key] += derivative
        derivatives_3[key] = initial[key] + derivative * dt
        derivatives_3[dx[i]] += dt

    for i, eq in enumerate(equations):
        derivative, key = eq(derivatives_3)
        k4[key] += derivative
        derivatives_4[key] = initial[key] + derivative * dt
        new_values[i] = initial[key] + (1 / 6) * dt * (k1[key] + 2 * k2[key] + 2 * k3[key] + k4[key])
    return new_values


ic.enable()
V_sound = v_sound(R - Rb)
mach=V/V_sound

cToDeg = 180 / m.pi
cToRad = m.pi / 180

tetta *= cToRad
initial = {}
dx = ['V', 'L', 'tetta', 'R']
equations = [dV_func, dL_func, dtetta_func, dR_func]

while mach > 1.32:

    tmich=t
    """этап 1 аэродинамическое торможение"""
    S, Cn, Fn, mass = 4.52, 0, 0, 1750
    V_sound = v_sound(R - Rb)
    ro = Get_ro(R - Rb)
    Cxa = Cx(V, V_sound)
    Px = mass / Cxa * S

    initial.update(
        {'S': S, 'Cn': Cn, 'Fn': Fn, 'tetta': tetta, 'Cxa': Cxa, 'ro': ro, 'L': L, 'V': V, 'R': R, 'mass': mass})
    values = runge_kutta_4(equations, initial, dt, dx)
    V = values[0]
    L = values[1]
    tetta = values[2]
    R = values[3]
    t += dt

    mach=V/V_sound
    CX.append(Cxa)
    TETTA.append(tetta * cToDeg)
    X.append(L)
    Y.append(R-Rb)
    V_MOD.append(V)
    T.append(t)
    plotnost.append(ro)
    napor.append(0.5*ro*V**2)
    nx.append((0.5 * S * Cxa * ro * V ** 2)/(mass*((gravy_const*mass_planet)/R**2)))
    PX.append(Px)

V_sound = v_sound(R - Rb)
mach = V/V_sound
print(f'V = {V:.3f}, tetta = {tetta * cToDeg:.3f}, L = {L:.3f}, H = {(R-Rb):.3f}, Mach={mach:.3f}, {t}')

while mach > 0.74:
    """этап 2 спуск на паращюте увода """
    S, Cn, Fn, mass = 4.52, 0.65, 6, 1750
    V_sound = v_sound(R - Rb)
    ro = Get_ro(R - Rb)
    Cxa = Cx(V, V_sound)
    Px = mass / Cxa * S

    initial.update(
        {'S': S, 'Cn': Cn, 'Fn': Fn, 'tetta': tetta, 'Cxa': Cxa, 'ro': ro, 'L': L, 'V': V, 'R': R, 'mass': mass})
    values = runge_kutta_4(equations, initial, dt, dx)
    V = values[0]
    L = values[1]
    tetta = values[2]
    R = values[3]
    t += dt

    mach=V/V_sound
    CX.append(Cxa)
    TETTA.append(tetta * cToDeg)
    X.append(L)
    Y.append(R-Rb)
    V_MOD.append(V)
    T.append(t)
    plotnost.append(ro)
    napor.append(0.5*ro*V**2)
    nx.append((0.5 * S * Cxa * ro * V ** 2)/(mass*((gravy_const*mass_planet)/R**2)))
    PX.append(Px)

V_sound = v_sound(R - Rb)
mach=V/V_sound
print(f'V = {V:.3f}, tetta = {tetta * cToDeg:.3f}, L = {L:.3f}, H = {(R-Rb):.3f}, Mach={mach:.3f}, {t}')


while t <= 71: #было 70 по циклограмме
    """третий этап спуск с верхней полусферой на парашюте увода"""
    S, Cn, Fn, mass = 4.155, 0.65, 6, 375
    V_sound = v_sound(R - Rb)
    ro = Get_ro(R - Rb)
    Cxa = 1.28
    Px = mass / Cxa * S

    initial.update(
        {'S': S, 'Cn': Cn, 'Fn': Fn, 'tetta': tetta, 'Cxa': Cxa, 'ro': ro, 'L': L, 'V': V, 'R': R, 'mass': mass})
    values = runge_kutta_4(equations, initial, dt, dx)
    V = values[0]
    L = values[1]
    tetta = values[2]
    R = values[3]
    t += dt

    CX.append(Cxa)
    TETTA.append(tetta * cToDeg)
    X.append(L)
    Y.append(R-Rb)
    V_MOD.append(V)
    T.append(t)
    plotnost.append(ro)
    napor.append(0.5*ro*V**2)
    nx.append((0.5 * S * Cxa * ro * V ** 2)/(mass*((gravy_const*mass_planet)/R**2)))
    PX.append(Px)

V_sound = v_sound(R - Rb)
mach=V/V_sound
print(f'V = {V:.3f}, tetta = {tetta * cToDeg:.3f}, L = {L:.3f}, H = {(R-Rb):.3f}, Mach={mach:.3f}, {t}')


while t <= 231: #while mach > 0.14: # # было 220 по циклограмме
    """четвертый этап спуск на стабилизирующем парашюте"""
    S, Cn, Fn, mass = 2.895, 0.78, 1.5, 120
    V_sound = v_sound(R - Rb)
    ro = Get_ro(R - Rb)
    Cxa = 0.58
    Px = mass / Cxa * S

    initial.update(
        {'S': S, 'Cn': Cn, 'Fn': Fn, 'tetta': tetta, 'Cxa': Cxa, 'ro': ro, 'L': L, 'V': V, 'R': R, 'mass': mass})
    values = runge_kutta_4(equations, initial, dt, dx)
    V = values[0]
    L = values[1]
    tetta = values[2]
    R = values[3]
    t += dt

    CX.append(Cxa)
    TETTA.append(tetta * cToDeg)
    X.append(L)
    Y.append(R-Rb)
    V_MOD.append(V)
    T.append(t)
    plotnost.append(ro)
    napor.append(0.5*ro*V**2)
    nx.append((0.5 * S * Cxa * ro * V ** 2)/(mass*((gravy_const*mass_planet)/R**2)))
    PX.append(Px)
V_sound = v_sound(R - Rb)
mach=V/V_sound
print(f'V = {V:.3f}, tetta = {tetta * cToDeg:.3f}, L = {L:.3f}, H = {(R-Rb):.3f}, Mach={mach:.3f}, {t}')


while t <= 400: #while mach > 0.03: # было 400 по циклограмме
    """пятый этап спуск на парашюте ввода аэростата """
    S, Cn, Fn, mass= 2.895, 0.97, 35, 120
    V_sound = v_sound(R - Rb)
    ro = Get_ro(R - Rb)
    Cxa = 0.58
    Px = mass / Cxa * S

    initial.update(
        {'S': S, 'Cn': Cn, 'Fn': Fn, 'tetta': tetta, 'Cxa': Cxa, 'ro': ro, 'L': L, 'V': V, 'R': R, 'mass': mass})
    values = runge_kutta_4(equations, initial, dt, dx)
    V = values[0]
    L = values[1]
    tetta = values[2]
    R = values[3]
    t += dt

    CX.append(Cxa)
    TETTA.append(tetta * cToDeg)
    X.append(L)
    Y.append(R-Rb)
    V_MOD.append(V)
    T.append(t)
    plotnost.append(ro)
    napor.append(0.5*ro*V**2)
    nx.append((0.5 * S * Cxa * ro * V ** 2)/(mass*((gravy_const*mass_planet)/R**2)))
    PX.append(Px)

V_sound = v_sound(R - Rb)
mach=V/V_sound
print(f'V = {V:.3f}, tetta = {tetta * cToDeg:.3f}, L = {L:.3f}, H = {(R-Rb):.3f}, Mach = {mach:.3f}, t = {t}')

for i in range(1, len(V_MOD)):
    derivative_value = (V_MOD[i] - V_MOD[i - 1]) / dt
    acceleration.append(derivative_value)
"""дальше вывод графиков"""
plt.plot(X, Y)
plt.title('Траектория')
plt.xlabel('Дальность, м')
plt.ylabel('Высота, м')
plt.grid(True)
plt.show()

plt.plot(T, Y)
plt.title('Зависимость высоты от времени')
plt.xlabel('Время, м')
plt.ylabel('Высота, м')
plt.grid(True)
plt.show()

plt.plot(T, V_MOD)
plt.title('Зависимость модуля скорости от времени')
plt.xlabel("Время, c")
plt.ylabel('Модуль скорости, м/с')
plt.grid(True)
plt.show()

plt.plot(V_MOD, Y)
plt.title('Зависимость модуля скорости от высоты')
plt.xlabel("Модуль скорости, м/с")
plt.ylabel('Высота, м')
plt.grid(True)
plt.show()

'''plt.plot( plotnost, V_MOD)
plt.title('Зависимость модуля скорости от плотности')
plt.xlabel('Плотность кг/м^3')
plt.ylabel('Модуль скорости, м/с')
plt.grid(True)
plt.show()'''

plt.plot(T, TETTA)
plt.title('Зависимость угла входа от времени')
plt.xlabel('Время, c')
plt.ylabel('TETTA, град')
plt.grid(True)
plt.show()

plt.plot(Y, TETTA)
plt.title('Зависимость угла входа от высоты')
plt.xlabel('Высота, м')
plt.ylabel('TETTA, град')
plt.grid(True)
plt.show()

'''plt.plot(LAM, PHI)
plt.title('Траектория')
plt.xlabel('Широта, град')
plt.ylabel('Долгота, град')
plt.grid(True)
plt.show()'''

'''plt.plot(T, CX)
plt.title('Аэродинамические коэффициенты')
plt.xlabel('Время, с')
plt.ylabel('Сxa')
plt.grid(True)
plt.show()'''

plt.plot(T, napor)
plt.title('Зависимость коростного напора от времени')
plt.xlabel('Время, с')
plt.ylabel('Скоростной напор, Па')
plt.grid(True)
plt.show()

T.pop()
plt.plot(T, acceleration)
plt.title('Зависимость ускорения от времени')
plt.xlabel('Время, с')
plt.ylabel('Ускорение м/с^2')
plt.grid(True)
plt.show()

nx.pop()
plt.plot(T, nx)
plt.title('Зависимость перегрузки от времени')
plt.xlabel('Время, с')
plt.ylabel('Перегрузка, g')
plt.grid(True)
plt.show()

PX.pop()
plt.plot(T, PX)
plt.title('Зависимость давления на мидель от времени')
plt.xlabel('Время, с')
plt.ylabel('Px, кг/м^2')
plt.grid(True)
plt.show()

end_time = time.time()
elapsed_time = end_time - start_time
ic(elapsed_time)

wb = Workbook()  # Создаем новый объект Workbook от openpyxl
ws = wb.active  # Делаем активным первый (и единственный) лист в новой книге
step = 100
headers = ["PX", "Перегрузка", "Плотность", "Ускорение", "Скоростной напор", "Угол входа",
           "Коэффициент лобового сопротивления", "X", "Y", "Скорость", "Время"]  # Список заголовков для данных
for col_num, header in enumerate(headers, start=1):
    ws.cell(row=2, column=col_num, value=header)  # Записываем заголовки в строку, следующую за параметрами

    # Записываем данные массивов в столбцы с заданным шагом
max_len = max(len(PX), len(nx), len(plotnost), len(acceleration), len(napor), len(TETTA), len(CX), len(X), len(Y),
              len(V_MOD), len(T))  # Находим максимальную длину среди массивов данных
start_row = 3  # Начинаем с первой строки после заголовков

for row_num in range(0, max_len, step):
    if row_num < len(PX):
        ws.cell(row=start_row + row_num // step, column=1, value=PX[row_num])  # Записываем данные времени
    if row_num < len(nx):
        ws.cell(row=start_row + row_num // step, column=2, value=nx[row_num])  # Записываем данные a
    if row_num < len(plotnost):
        ws.cell(row=start_row + row_num // step, column=3, value=plotnost[row_num])  # Записываем данные p
    if row_num < len(acceleration):
        ws.cell(row=start_row + row_num // step, column=4, value=acceleration[row_num])  # Записываем данные e
    if row_num < len(napor):
        ws.cell(row=start_row + row_num // step, column=5, value=napor[row_num])  # Записываем данные r
    if row_num < len(TETTA):
        ws.cell(row=start_row + row_num // step, column=6, value=TETTA[row_num])  # Записываем данные omega
    if row_num < len(CX):
        ws.cell(row=start_row + row_num // step, column=7, value=CX[row_num])
    if row_num < len(X):
        ws.cell(row=start_row + row_num // step, column=8, value=X[row_num])
    if row_num < len(Y):
        ws.cell(row=start_row + row_num // step, column=9, value=Y[row_num])
    if row_num < len(V_MOD):
        ws.cell(row=start_row + row_num // step, column=10, value=V_MOD[row_num])
    if row_num < len(T):
        ws.cell(row=start_row + row_num // step, column=11, value=T[row_num])
# Выбираем путь для сохранения файла
file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

if file_path:
    wb.save(file_path)  # Если путь выбран, сохраняем файл по указанному пути
